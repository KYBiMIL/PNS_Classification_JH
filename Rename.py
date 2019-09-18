import os
import sys
import numpy as np
import cv2
import glob
from PIL import Image
import PIL
import pandas as pd
import shutil
import xlwt
import openpyxl



'''
Coded by Juhui
on 2019.08.23
'''



'''
Wide_AMD_0001_R
Wide_DMR_0001_R
Wide_NOR_0001_R
Wide_RVO_0001_R

General_Glaucoma_0001_R
General_Normal_0001_R
'''


# fn	'1차 메타데이터.xlsx'
# sn	'일반_정상'
#	'일반_녹내장'
#	'광각_정상'
#	'광각_AMD'
#	'광각_DMR'
#	'광각_RVO'



def Make_Directory(path):

	if not os.path.exists(path):
		os.makedirs(path)

	return True



def SheetAndName(sheetname):

	if sheetname == '일반_정상':
		add_ID = 'General_Normal_'
		imgDir = '/data/Fundus Image/g_normal'
	elif sheetname == '일반_녹내장':
		add_ID = 'General_Glaucoma_'
		imgDir = '/data/Fundus Image/g_glaucoma'
	elif sheetname == '광각_정상':
		add_ID = 'Wide_NOR_'
		imgDir = '/data/Fundus Image/w_normal'
	elif sheetname == '광각_AMD':
		add_ID = 'Wide_AMD_'
		imgDir = '/data/Fundus Image/w_amd'
	elif sheetname == '광각_DMR':
		add_ID = 'Wide_DMR_'
		imgDir = '/data/Fundus Image/w_dmr'
	elif sheetname == '광각_RVO':
		add_ID = 'Wide_RVO_'
		imgDir = '/data/Fundus Image/w_rvo'

	return add_ID, imgDir


def Indexing2Four(num):

	n = str(num)

	if len(n) == 1:
		n = '000' + n
	elif len(n) == 2:
		n = '00' + n
	elif len(n) == 3:
		n = '0' + n
	elif len(n) == 4:
		n = n
	else:
		print("error!!\n")

	return n


def Direction_of_K2E(direction):

	if direction == '좌' or direction == 'L':
		res = 'L'
	else:
		res = 'R'

	return res



def RenameJPG(path, fn, sn, xlFileDir):	# = '/data/Fundus Image', fn, sn):

	df = pd.read_excel(os.path.join(path, fn), sheetname = sn)
	d_info = df['이미지화일명'], df['좌/우']

	Make_Directory(os.path.join(path, 'AnonymizedData'))
	

	if not os.path.isfile(xlFileDir):
		wb = openpyxl.Workbook()
		shit = wb.active
		wb.save(xlFileDir)
	
	wb = openpyxl.load_workbook(xlFileDir)
	sht = wb.create_sheet(sn)
	sht.cell(row=1, column=1).value = 'oldName'
	sht.cell(row=1, column=2).value = 'newName'	

	
	add_ID, imgDir = SheetAndName(sn)

	for i in range(len(df)):
		oldID = d_info[0][i]
		Direction = Direction_of_K2E(d_info[1][i])
		newID = add_ID + Indexing2Four(i + 1) + '_' + Direction

		imgPath = os.path.join(imgDir, oldID + '.jpg')
		newPath = os.path.join(path, 'AnonymizedData', imgDir.split('/')[-1], newID + '.jpg')
		Make_Directory(os.path.dirname(newPath))

		shutil.copy2(imgPath, newPath)

		sht.cell(row=i+2, column=1).value = oldID
		sht.cell(row=i+2, column=2).value = newID
		
	wb.save(xlFileDir)		


	return True



path = '/data/Fundus Image'
xlFileDir = os.path.join(path, 'AnonymizedData', 'ID_Matching.xlsx')

RenameJPG(path = path, fn = '1차 메타데이터_REV.xlsx', sn = '일반_녹내장', xlFileDir = xlFileDir)


'''
## sn

일반_정상
일반_녹내장
광각_정상
광각_AMD
광각_DMR
광각_RVO
'''	



		

	










